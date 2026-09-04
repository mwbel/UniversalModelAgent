#!/usr/bin/env python3
"""Export canonical MATLAB workbook outputs to the web app's JSON oracle."""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

from openpyxl import load_workbook


YEAR_FILE_PATTERN = re.compile(r"^(\d{4})年时轮历-check\.xlsx$")


def read_cell(value):
    return int(value) if isinstance(value, (int, float)) and value == int(value) else value


def read_month(sheet) -> dict:
    rows = list(sheet.iter_rows(min_row=1, max_row=sheet.max_row, max_col=10, values_only=True))

    def value(row: int, column: int):
        return rows[row - 1][column - 1]

    summary = {
        "积月闰余": [read_cell(value(7, 6)), read_cell(value(8, 6))],
        "曜基数": [read_cell(value(row, 7)) for row in range(7, 12)],
        "整零数": [read_cell(value(row, 9)) for row in range(7, 9)],
        "太阳基数": [read_cell(value(row, 8)) for row in range(7, 12)],
    }

    days = []
    for day in range(1, 31):
        start_row = 8 * day + 6
        days.append(
            {
                "day": day,
                "定曜": [read_cell(value(start_row + offset, 7)) for offset in range(6)],
                "月伴星宿": [read_cell(value(start_row + offset, 8)) for offset in range(6)],
                "定日": [read_cell(value(start_row + offset, 9)) for offset in range(5)],
                "会合": [read_cell(value(start_row + offset, 10)) for offset in range(6)],
            }
        )
    return {"summary": summary, "days": days}


def export(output_dir: Path, output_path: Path) -> None:
    years = {}
    for workbook_path in sorted(output_dir.glob("*年时轮历-check.xlsx")):
        match = YEAR_FILE_PATTERN.match(workbook_path.name)
        if match is None:
            continue
        year = int(match.group(1))
        workbook = load_workbook(workbook_path, data_only=True, read_only=True)
        months = {}
        excluded_months = []
        for sheet_name in workbook.sheetnames:
            if not sheet_name.isdigit() or not 1 <= int(sheet_name) <= 12:
                continue
            month = int(sheet_name)
            if workbook[sheet_name].max_row < 251:
                excluded_months.append(month)
                continue
            months[month] = read_month(workbook[sheet_name])
        if not months:
            raise ValueError(f"{workbook_path.name} has no complete monthly sheets")
        years[str(year)] = {
            "sourceFile": workbook_path.name,
            "sourceMode": "current_local",
            "excludedMonths": excluded_months,
            "months": {str(month): months[month] for month in sorted(months)},
        }
        workbook.close()

    if not years:
        raise FileNotFoundError(f"No canonical MATLAB workbooks found in {output_dir}")

    payload = {
        "schemaVersion": 1,
        "sourcePolicy": "程序-时轮历/输出/*年时轮历-check.xlsx; check0 and non-check variants excluded",
        "years": years,
    }
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def main() -> int:
    project_root = Path(__file__).resolve().parents[3]
    parser = argparse.ArgumentParser()
    parser.add_argument("--output-dir", type=Path, default=project_root / "程序-时轮历" / "输出")
    parser.add_argument("--output", type=Path, default=Path(__file__).resolve().parents[1] / "data" / "matlab_oracle.json")
    args = parser.parse_args()
    export(args.output_dir, args.output)
    print(f"exported {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
